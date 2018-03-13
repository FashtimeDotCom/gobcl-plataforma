from openpyxl import load_workbook

from .load_data.ministries import create_ministry
from .load_data.public_enterprises import create_public_enterprise
from .load_data.campaigns import create_campaign
from .load_data.regions import create_region
from .load_data.public_servants import create_public_servant
from .load_data.communes import create_commune
from .load_data.public_services import create_public_service

from government_structures.models import GovernmentStructure


def create_commune_data(government_structure=None):

    if not government_structure:
        government_structure = GovernmentStructure.objects.get_or_none(
            current_government=True)

    wb = load_workbook('catastro.xlsx')

    for sheet_name in wb.get_sheet_names():
        if not sheet_name == 'comunas':
            continue

        sheet = wb.get_sheet_by_name(sheet_name)

        for row in sheet.rows:
            create_commune(row, government_structure)


def load_data_from_xlsx():
    wb = load_workbook('gobcl.xlsx')

    sheets = {
        'Ministerios': create_ministry,
        'Subsecretarios': create_public_servant,
        'Intendencias': create_region,
        'Servicios Públicos': create_public_service,
        'Empresas Públicas': create_public_enterprise,
        'Campañas': create_campaign,
        'Acerca de Chile': None,
        'Nuestro país': None,
        'Presidencia': None,
    }

    government_structure = GovernmentStructure.objects.get_or_none(
        current_government=True)

    for sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_name)

        for row in sheet.rows:
            if sheets.get(sheet_name):
                sheets.get(sheet_name)(row, government_structure)

    create_commune_data(government_structure)
